"""Export EventFeedbackResponse querysets to Excel (.xlsx) and PDF."""

from io import BytesIO
from datetime import datetime

from django.http import HttpResponse
from django.utils import timezone


# (field_name, column header)
EXPORT_COLUMNS = [
    ('id', 'ID'),
    ('campaign_title', 'Campaign / Event'),
    ('name', 'Name'),
    ('visitor_type', 'Visitor type'),
    ('institution_name', 'Institution'),
    ('city_village', 'City / Village'),
    ('attractions', 'What attracted you'),
    ('heard_from', 'How did you hear'),
    ('overall_rating', 'Overall rating'),
    ('organization_rating', 'Organization'),
    ('hospitality_rating', 'Hospitality'),
    ('atmosphere_rating', 'Atmosphere'),
    ('stage_programme_rating', 'Stage & programme'),
    ('crowd_management_rating', 'Crowd management'),
    ('facilities_rating', 'Facilities'),
    ('attended_meet_greet', 'Attended Meet & Greet'),
    ('meet_greet_rating', 'Meet & Greet rating'),
    ('excitement_level', 'Excitement level'),
    ('presence_made_exciting', 'Presence made exciting'),
    ('enjoy_most_meet_greet', 'Enjoyed most (Meet & Greet)'),
    ('college_knowledge', 'College knowledge'),
    ('learned_experienced', 'Learned / experienced'),
    ('campus_impression', 'Campus impression'),
    ('contribution_areas', 'Contribution areas'),
    ('contribution_other_suggestion', 'Contribution other suggestion'),
    ('memorable_scale', 'Memorable (1–5)'),
    ('attend_future_scale', 'Attend future (1–5)'),
    ('recommend_events_scale', 'Recommend events (1–5)'),
    ('another_celebrity_meet', 'Another celebrity Meet & Greet'),
    ('best_part', 'Best part'),
    ('improvements', 'Improvements'),
    ('message_for_guest', 'Message for guest'),
    ('additional_comments', 'Additional comments'),
    ('final_description', 'Final description'),
    ('submitted_at', 'Submitted at'),
]


def _cell_value(obj, field_name):
    if field_name == 'campaign_title':
        return obj.campaign.title if obj.campaign_id else ''
    val = getattr(obj, field_name, '')
    if field_name == 'submitted_at' and val:
        if timezone.is_aware(val):
            val = timezone.localtime(val)
        return val.strftime('%Y-%m-%d %H:%M:%S')
    if val is None:
        return ''
    return val


def rows_for_export(queryset):
    """Yield list of cell values in EXPORT_COLUMNS order."""
    qs = queryset.select_related('campaign').order_by('-submitted_at')
    for obj in qs:
        yield [_cell_value(obj, name) for name, _ in EXPORT_COLUMNS]


def export_filename(prefix, ext):
    stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    return f'{prefix}_{stamp}.{ext}'


def build_excel_response(queryset, filename_prefix='event_feedback_responses'):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'Responses'

    headers = [h for _, h in EXPORT_COLUMNS]
    ws.append(headers)

    header_fill = PatternFill('solid', fgColor='E61013')
    header_font = Font(bold=True, color='FFFFFF')
    thin = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )
    for col, _ in enumerate(headers, start=1):
        cell = ws.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical='center')

    count = 0
    for row in rows_for_export(queryset):
        ws.append(row)
        count += 1
        for col in range(1, len(headers) + 1):
            c = ws.cell(ws.max_row, col)
            c.alignment = Alignment(wrap_text=True, vertical='top')
            c.border = thin

    # Reasonable column widths
    widths = {
        1: 8, 2: 28, 3: 18, 4: 16, 5: 22, 6: 14,
    }
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = widths.get(col, 16)

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = 'A2'

    # Summary sheet
    summary = wb.create_sheet('Summary', 0)
    summary['A1'] = 'Event Feedback Responses – Export'
    summary['A1'].font = Font(bold=True, size=14, color='E61013')
    summary['A2'] = 'Generated at'
    summary['B2'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    summary['A3'] = 'Total responses'
    summary['B3'] = count
    summary.column_dimensions['A'].width = 22
    summary.column_dimensions['B'].width = 40

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = export_filename(filename_prefix, 'xlsx')
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def build_pdf_response(queryset, filename_prefix='event_feedback_responses'):
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate,
        Paragraph,
        Spacer,
        Table,
        TableStyle,
        PageBreak,
        KeepTogether,
    )

    buffer = BytesIO()
    page = landscape(A4)
    doc = SimpleDocTemplate(
        buffer,
        pagesize=page,
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title='Event Feedback Responses',
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'TitleRed',
        parent=styles['Heading1'],
        textColor=colors.HexColor('#E61013'),
        fontSize=16,
        spaceAfter=6,
        alignment=TA_CENTER,
    )
    meta_style = ParagraphStyle(
        'Meta',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#555555'),
        alignment=TA_CENTER,
        spaceAfter=12,
    )
    heading_style = ParagraphStyle(
        'CardHead',
        parent=styles['Heading3'],
        fontSize=11,
        textColor=colors.HexColor('#E61013'),
        spaceBefore=4,
        spaceAfter=4,
    )
    body_style = ParagraphStyle(
        'CardBody',
        parent=styles['Normal'],
        fontSize=8,
        leading=11,
        alignment=TA_LEFT,
    )
    label_style = ParagraphStyle(
        'Label',
        parent=styles['Normal'],
        fontSize=7,
        textColor=colors.HexColor('#666666'),
        leading=9,
    )
    value_style = ParagraphStyle(
        'Value',
        parent=styles['Normal'],
        fontSize=8,
        leading=10,
    )

    def p(text, style=value_style):
        text = '' if text is None else str(text)
        text = (
            text.replace('&', '&amp;')
            .replace('<', '&lt;')
            .replace('>', '&gt;')
            .replace('\n', '<br/>')
        )
        return Paragraph(text or '—', style)

    qs = list(queryset.select_related('campaign').order_by('-submitted_at'))
    story = [
        Paragraph('Event Feedback Responses', title_style),
        Paragraph(
            f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} &nbsp;|&nbsp; '
            f'Total: {len(qs)}',
            meta_style,
        ),
    ]

    # Compact overview table (key columns)
    overview_headers = [
        'ID', 'Name', 'Campaign', 'Type', 'Overall', 'Final', 'Submitted',
    ]
    overview_data = [[Paragraph(h, label_style) for h in overview_headers]]
    for obj in qs:
        submitted = _cell_value(obj, 'submitted_at')
        overview_data.append([
            p(obj.pk, value_style),
            p(obj.name, value_style),
            p(obj.campaign.title if obj.campaign_id else '', value_style),
            p(obj.visitor_type, value_style),
            p(obj.overall_rating, value_style),
            p(obj.final_description, value_style),
            p(submitted, value_style),
        ])

    overview = Table(
        overview_data,
        colWidths=[18 * mm, 35 * mm, 55 * mm, 35 * mm, 18 * mm, 40 * mm, 35 * mm],
        repeatRows=1,
    )
    overview.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E61013')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FFF5F5')]),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#CCCCCC')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    story.append(overview)
    story.append(Spacer(1, 8 * mm))
    story.append(Paragraph('Detailed responses', heading_style))
    story.append(Spacer(1, 3 * mm))

    # Detail cards (all fields) – two-column label/value tables
    detail_fields = [(n, h) for n, h in EXPORT_COLUMNS if n not in ('id',)]
    for idx, obj in enumerate(qs):
        block = [
            Paragraph(
                f'#{obj.pk} — {obj.name} ({_cell_value(obj, "submitted_at")})',
                heading_style,
            )
        ]
        rows = []
        for field_name, header in detail_fields:
            rows.append([
                p(header, label_style),
                p(_cell_value(obj, field_name), value_style),
            ])
        t = Table(rows, colWidths=[55 * mm, 200 * mm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#F8F9FA')),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#DDDDDD')),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ]))
        block.append(t)
        block.append(Spacer(1, 6 * mm))
        story.append(KeepTogether(block))
        if idx < len(qs) - 1 and (idx + 1) % 2 == 0:
            story.append(PageBreak())

    if not qs:
        story.append(Paragraph('No responses to export.', body_style))

    doc.build(story)
    buffer.seek(0)

    filename = export_filename(filename_prefix, 'pdf')
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
